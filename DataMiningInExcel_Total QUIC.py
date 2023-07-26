import pandas as pd
import numpy as np
from openpyxl import load_workbook
from operator import itemgetter
import statistics
import time



start_time=time.time()

result_file="/Users/guqingxian/Desktop/学习/Data Mining in Excel/Results_Exercises/results.xlsx"
StandAlone=pd.read_excel("/Users/guqingxian/Desktop/学习/Data Mining in Excel/Data Set/Stand-Alone.xlsx")
book=load_workbook('/Users/guqingxian/Desktop/学习/Data Mining in Excel/Data Set/Stand-Alone.xlsx')
datasheet=book['单机 all30']
resultbook=load_workbook(result_file)
resultsheet=resultbook['Total QUIC']

Row=0
def findNo(column_index, target_string, trace, sheet):
    # trace:第n个出现的位置，例如 1 表示第一个出现的位置
    # column index: 列的数字索引，例如 3 表示第三列
    # 遍历指定列，找到第n个出现的目标字符串
    result = None
    count = 0
    # global startrow
    global Row

    # print('startrow=', startrow)
    for row in sheet.iter_rows(min_row=Row + 1, min_col=column_index, max_col=column_index, values_only=True):
        cell_value = row[0]
        Row += 1
        if cell_value == target_string:
            count += 1

            if count == trace:
                result = Row
                # print('Row=', Row)
                # startrow = Row+1
                # print('startrow==',startrow)
                break
    if result is not None:
        return Row
    else:
        return sheet.max_row - 1


traces=[]
for i in range(30):
    traces.append(findNo(1,'No.',1,datasheet))

Row=0

QUIC_StandAlone_Delay=[0] * 30
QUIC_StandAlone_Length=[0] * 30
cnt=1
for row in datasheet.iter_rows(min_row=2,min_col=2, max_col=6, values_only=True):
    if row[3] == 'QUIC' :
        for i in range(29):
            if cnt >= traces[i] and cnt <= traces[i+1] :
                QUIC_StandAlone_Delay[i] += row[0]
                QUIC_StandAlone_Length[i] += row[4]
                print(cnt)
                break
            elif cnt >= traces[29] :
                QUIC_StandAlone_Delay[29] += row[0]
                QUIC_StandAlone_Length[29] += row[4]
    cnt+=1

cnt=1
for i in range(30):
    resultsheet.cell(row=i+3,column=2,value=QUIC_StandAlone_Length[i])
    resultsheet.cell(row=i+3,column=7,value=QUIC_StandAlone_Delay[i])


resultsheet.cell(row=33,column=2,value=statistics.median(QUIC_StandAlone_Length))
resultsheet.cell(row=34,column=2,value=statistics.mean(QUIC_StandAlone_Length))
resultsheet.cell(row=33,column=7,value=statistics.median(QUIC_StandAlone_Delay))
resultsheet.cell(row=34,column=7,value=statistics.mean(QUIC_StandAlone_Delay))

print('before 1v1 load_workbook')
book_1v1=load_workbook('/Users/guqingxian/Desktop/学习/Data Mining in Excel/Data Set/1v1.xlsx')
datasheet_1v1=book_1v1['backup']
print('after 1v1 load_wordbook')
traces_1v1=[]
for i in range(30):
    traces_1v1.append(findNo(2,1,1,datasheet_1v1))
Row=0

QUIC_1v1_Delay=[0] * 30
QUIC_1v1_Length=[0] * 30
cnt=1
for row in datasheet_1v1.iter_rows(min_row=2,min_col=5, max_col=9, values_only=True):
    if row[3] == 'QUIC' :
        for i in range(29):
            if cnt >= traces_1v1[i] and cnt <= traces_1v1[i+1] :
                QUIC_1v1_Delay[i] += row[0]
                QUIC_1v1_Length[i] += row[4]
                print(cnt)
                break
            elif cnt >= traces_1v1[29] :
                QUIC_1v1_Delay[29] += row[0]
                QUIC_1v1_Length[29] += row[4]
    cnt+=1

cnt=1
for i in range(30):
    resultsheet.cell(row=i+3,column=3,value=QUIC_1v1_Length[i])
    resultsheet.cell(row=i+3,column=8,value=QUIC_1v1_Delay[i])


resultsheet.cell(row=33,column=3,value=statistics.median(QUIC_1v1_Length))
resultsheet.cell(row=34,column=3,value=statistics.mean(QUIC_1v1_Length))
resultsheet.cell(row=33,column=8,value=statistics.median(QUIC_1v1_Delay))
resultsheet.cell(row=34,column=8,value=statistics.mean(QUIC_1v1_Delay))

print('before loading 5v5 workbook')
book_5v5=load_workbook('/Users/guqingxian/Desktop/学习/Data Mining in Excel/Data Set/5v5.xlsx')
datasheet_5v5=book_5v5['5v5']
print('after loading 5v5 workbook')

traces_5v5=[]
for i in range(30):
    traces_5v5.append(findNo(2,1,1,datasheet_5v5))
Row=0

QUIC_5v5_Delay=[0] * 30
QUIC_5v5_Length=[0] * 30
cnt=1
for row in datasheet_5v5.iter_rows(min_row=2,min_col=5, max_col=9, values_only=True):
    if row[3] == 'QUIC' :
        for i in range(29):
            if cnt >= traces_5v5[i] and cnt <= traces_5v5[i+1] :
                QUIC_5v5_Delay[i] += row[0]
                QUIC_5v5_Length[i] += row[4]
                print(cnt)
                break
            elif cnt >= traces_5v5[29] :
                QUIC_5v5_Delay[29] += row[0]
                QUIC_5v5_Length[29] += row[4]

    cnt+=1

cnt=1
for i in range(30):
    resultsheet.cell(row=i+3,column=4,value=QUIC_5v5_Length[i])
    resultsheet.cell(row=i+3,column=9,value=QUIC_5v5_Delay[i])


resultsheet.cell(row=33,column=4,value=statistics.median(QUIC_5v5_Length))
resultsheet.cell(row=34,column=4,value=statistics.mean(QUIC_5v5_Length))
resultsheet.cell(row=33,column=9,value=statistics.median(QUIC_5v5_Delay))
resultsheet.cell(row=34,column=9,value=statistics.mean(QUIC_5v5_Delay))



resultbook.save(result_file)
resultbook.close()
book.close()
book_1v1.close()
book_5v5.close()

end_time=time.time()
print('total time is ',end_time-start_time)