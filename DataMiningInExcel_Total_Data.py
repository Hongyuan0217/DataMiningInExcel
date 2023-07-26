import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl import Workbook
import time

start_time=time.time()


result_file="/Users/guqingxian/Desktop/学习/Data Mining in Excel/Results_Exercises/results.xlsx"
book = load_workbook(result_file)
sheet=book['Overall to Total']
# xls = pd.ExcelFile(result_file)
df=pd.DataFrame(columns=['Max','Min','Avg','Sum','Package Count'])
excelWriter = pd.ExcelWriter(result_file, engine='openpyxl', mode='a', if_sheet_exists='new')


# result=pd.read_excel(result_file,sheet_name='Overall to Total',header=None,engine='openpyxl')
row=3

def EnterDataInExcel(name):
    global row, df
    Max = name["Length"].max()
    sheet.cell(row=row,column=3,value=Max)
    Min =name["Length"].min()
    sheet.cell(row=row, column=4, value=Min)
    Avg = name['Length'].mean()
    sheet.cell(row=row, column=5, value=round(Avg))
    Sum = name['Length'].sum()
    sheet.cell(row=row, column=6, value=Sum)
    PackageCount = name['Length'].shape[0]
    sheet.cell(row=row, column=7, value=PackageCount)
    # df=df._append({'Max':Max,'Min':Min,'Avg':Avg,'Sum':Sum,'PackageCount':PackageCount},ignore_index=True)
    row += 1

def AllTypesOfData(name):
    StandAlone = pd.read_excel(
        "/Users/guqingxian/Desktop/学习/Data Mining in Excel/Data Set/Stand-Alone - "+"{}".format(name)+".xlsx")
    A1v1 = pd.read_excel("/Users/guqingxian/Desktop/学习/Data Mining in Excel/Data Set/1v1 - "+"{}".format(name)+".xlsx")
    A5v5 = pd.read_excel("/Users/guqingxian/Desktop/学习/Data Mining in Excel/Data Set/5v5 - "+"{}".format(name)+".xlsx")
    EnterDataInExcel(StandAlone)
    EnterDataInExcel(A1v1)
    EnterDataInExcel(A5v5)

AllTypesOfData('HTTP')
AllTypesOfData('DNS')
AllTypesOfData('MDNS')
AllTypesOfData('LLMNR')
AllTypesOfData('NBNS')
AllTypesOfData("SSDP")
AllTypesOfData('NTP')
AllTypesOfData("SSL")
AllTypesOfData('TLS')


#TCP is special because the sum of Length of the first sheet is different from the second one,
#so we should use the second sheet - same as the Result.xlsx
StandAlone = pd.read_excel(
    "/Users/guqingxian/Desktop/学习/Data Mining in Excel/Data Set/Stand-Alone - TCP.xlsx",sheet_name='All 30')
A1v1 = pd.read_excel("/Users/guqingxian/Desktop/学习/Data Mining in Excel/Data Set/1v1 - TCP.xlsx")
A5v5 = pd.read_excel("/Users/guqingxian/Desktop/学习/Data Mining in Excel/Data Set/5v5 - TCP.xlsx")
EnterDataInExcel(StandAlone)
EnterDataInExcel(A1v1)
EnterDataInExcel(A5v5)



AllTypesOfData('UDP')



#QUIC is special because it does not have StandAlone
QUIC_1v1=pd.read_excel("/Users/guqingxian/Desktop/学习/Data Mining in Excel/Data Set/1v1 - QUIC.xlsx")
QUIC_5v5=pd.read_excel("/Users/guqingxian/Desktop/学习/Data Mining in Excel/Data Set/5v5 - QUIC.xlsx")
sheet.cell(row=row,column=3,value=0)
sheet.cell(row=row,column=4,value=0)
sheet.cell(row=row,column=5,value=0)
sheet.cell(row=row,column=6,value=0)
sheet.cell(row=row,column=7,value=0)
row+=1
EnterDataInExcel(QUIC_1v1)
EnterDataInExcel(QUIC_5v5)

sheet.cell(row=row,column=5,value='Total')
sum=0
for i in range(row-3):
    cell=sheet.cell(row=i+3,column=6).value
    sum=sum+int(cell)

sheet.cell(row=row,column=6,value=sum)

sum=0
for i in range(row-3):
    cell=sheet.cell(row=i+3,column=7).value
    sum=sum+int(cell)
sheet.cell(row=row,column=7,value=sum)



book.save(result_file)

end_time=time.time()
total_time=end_time-start_time
print('total time is ',total_time)



# df.to_excel(excelWriter,index=False,startrow=2, startcol=2, header=False,sheet_name='Overall to Total')
# excelWriter._save()
# excelWriter.close()
# book.save()
# #some detailed format
# book = load_workbook(result_file);
# sheet=book['Overall to Total'];
#
# sheet.merge_cells('A1:A2')
# sheet.merge_cells('B1:B2')
# sheet.merge_cells('C1:F1')
# sheet['C1']="Packet Size (Bytes)"
# sheet['C1'].alignment = Alignment(horizontal='center', vertical='center')
# sheet.merge_cells('G1:G2')
# sheet['G1']='Package ' \
#             'Count'
# sheet.column_dimensions['B'].width = 18
# sheet.column_dimensions['G'].width = 18
# sheet['G1'].alignment = Alignment(horizontal='center', vertical='center')
#
# #save
# book.save(result_file);

