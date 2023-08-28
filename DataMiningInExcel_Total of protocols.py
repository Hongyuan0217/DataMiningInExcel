import pandas as pd
import numpy as np
from openpyxl import load_workbook
from operator import itemgetter
import statistics
import time



start_time=time.time()

result_file="/Users/guqingxian/Desktop/学习/Data Mining in Excel/Results_Exercises/results.xlsx"
StandAlone=pd.read_excel("/Users/guqingxian/Desktop/学习/Data Mining in Excel/Data Set/Stand-Alone.xlsx")
book=load_workbook('/Users/guqingxian/Desktop/学习/Data Mining in Excel/Data Set/Stand-Alone.xlsx')
datasheet=book['单机 all30']
resultbook=load_workbook(result_file)
# resultsheet=resultbook['Total QUIC']
Endtime_StandAlone=time.time()
loadingtime_StandAlone=Endtime_StandAlone-start_time

print('before 1v1 load_workbook')
book_1v1=load_workbook('/Users/guqingxian/Desktop/学习/Data Mining in Excel/Data Set/1v1.xlsx')
datasheet_1v1=book_1v1['backup']
print('after 1v1 load_wordbook')
Endtime_1v1=time.time()
loadingtime_1v1=Endtime_1v1-Endtime_StandAlone



print('before loading 5v5 workbook')
book_5v5=load_workbook('/Users/guqingxian/Desktop/学习/Data Mining in Excel/Data Set/5v5.xlsx')
datasheet_5v5=book_5v5['5v5']
print('after loading 5v5 workbook')
Endtime_5v5=time.time()
loadingtime_5v5=Endtime_5v5-Endtime_1v1
Row=0

def findNo(column_index, target_string, trace, sheet):
    # trace:第n个出现的位置，例如 1 表示第一个出现的位置
    # column index: 列的数字索引，例如 3 表示第三列
    # 遍历指定列，找到第n个出现的目标字符串
    result = None
    count = 0
    # global startrow
    global Row

    # print('startrow=', startrow)
    for row in sheet.iter_rows(min_row=Row + 1, min_col=column_index, max_col=column_index, values_only=True):
        cell_value = row[0]
        Row += 1
        if cell_value == target_string:
            count += 1

            if count == trace:
                result = Row
                # print('Row=', Row)
                # startrow = Row+1
                # print('startrow==',startrow)
                break
    if result is not None:
        return Row
    else:
        return sheet.max_row - 1


def getTotal_StandAlone(protocol):

    starttime=time.time()
    result_sheet_name='Total {}'.format(protocol)
    resultsheet=resultbook[result_sheet_name]
    global Row
    traces = []
    for i in range(30):
        traces.append(findNo(1, 'No.', 1, datasheet))

    Row = 0

    Delay = [0] * 30
    Length = [0] * 30
    cnt = 1
    packagecnt=0
    for row in datasheet.iter_rows(min_row=2, min_col=2, max_col=6, values_only=True):
        if row[3] == '{}'.format(protocol):
            packagecnt += 1
            for i in range(29):
                if cnt >= traces[i] and cnt <= traces[i + 1]:
                    Delay[i] += row[0]
                    Length[i] += row[4]
                    # print(cnt)
                    break
                elif cnt >= traces[29]:
                    Delay[29] += row[0]
                    Length[29] += row[4]
        cnt += 1

    print('there are ',packagecnt,' of packages in StandAlone {} '.format(protocol))
    for i in range(30):
        resultsheet.cell(row=i + 3, column=2, value=Length[i])
        resultsheet.cell(row=i + 3, column=7, value=Delay[i])

    resultsheet.cell(row=33, column=2, value=statistics.median(Length))
    resultsheet.cell(row=34, column=2, value=statistics.mean(Length))
    resultsheet.cell(row=33, column=7, value=statistics.median(Delay))
    resultsheet.cell(row=34, column=7, value=statistics.mean(Delay))
    endtime=time.time()
    print('The time of stand-alone {} is '.format(protocol),endtime-starttime+loadingtime_StandAlone)


def getTotal_1v1(protocol):

    starttime=time.time()
    result_sheet_name='Total {}'.format(protocol)
    resultsheet=resultbook[result_sheet_name]
    global Row
    traces_1v1 = []
    for i in range(30):
        traces_1v1.append(findNo(2, 1, 1, datasheet_1v1))
    Row = 0

    Delay = [0] * 30
    Length = [0] * 30
    cnt = 1
    packagecnt=0
    for row in datasheet_1v1.iter_rows(min_row=2, min_col=5, max_col=9, values_only=True):
        if row[3] == '{}'.format(protocol):
            packagecnt += 1
            for i in range(29):
                if cnt >= traces_1v1[i] and cnt <= traces_1v1[i + 1]:
                    Delay[i] += row[0]
                    Length[i] += row[4]
                    # print(cnt)
                    break
                elif cnt >= traces_1v1[29]:
                    Delay[29] += row[0]
                    Length[29] += row[4]
        cnt += 1

    print('there are ',packagecnt,' of packages in 1v1 {} '.format(protocol))
    for i in range(30):
        resultsheet.cell(row=i + 3, column=3, value=Length[i])
        resultsheet.cell(row=i + 3, column=8, value=Delay[i])

    resultsheet.cell(row=33, column=3, value=statistics.median(Length))
    resultsheet.cell(row=34, column=3, value=statistics.mean(Length))
    resultsheet.cell(row=33, column=8, value=statistics.median(Delay))
    resultsheet.cell(row=34, column=8, value=statistics.mean(Delay))
    endtime=time.time()
    print('The time of 1v1 {} is '.format(protocol),endtime-starttime+loadingtime_1v1)


def getTotal_5v5(protocol):
    starttime=time.time()
    result_sheet_name='Total {}'.format(protocol)
    resultsheet=resultbook[result_sheet_name]
    global Row
    traces_5v5 = []
    for i in range(30):
        traces_5v5.append(findNo(2, 1, 1, datasheet_5v5))
    Row = 0

    Delay = [0] * 30
    Length = [0] * 30
    cnt = 1
    packagecnt = 0
    for row in datasheet_5v5.iter_rows(min_row=2, min_col=5, max_col=9, values_only=True):
        if row[3] == '{}'.format(protocol):
            packagecnt += 1
            for i in range(29):
                if cnt >= traces_5v5[i] and cnt <= traces_5v5[i + 1]:
                    Delay[i] += row[0]
                    Length[i] += row[4]
                    # print(cnt)
                    break
                elif cnt >= traces_5v5[29]:
                    Delay[29] += row[0]
                    Length[29] += row[4]

        cnt += 1

    print('there are ',packagecnt,' of packages in 5v5 {} '.format(protocol))

    for i in range(30):
        resultsheet.cell(row=i + 3, column=4, value=Length[i])
        resultsheet.cell(row=i + 3, column=9, value=Delay[i])

    resultsheet.cell(row=33, column=4, value=statistics.median(Length))
    resultsheet.cell(row=34, column=4, value=statistics.mean(Length))
    resultsheet.cell(row=33, column=9, value=statistics.median(Delay))
    resultsheet.cell(row=34, column=9, value=statistics.mean(Delay))
    endtime=time.time()
    print('The time of 5v5 {} is '.format(protocol),endtime-starttime+loadingtime_5v5)


protocols=['HTTP','DNS','MDNS','LLMNR','NBNS','SSDP','NTP','SSL','TLSv1.2','TCP','UDP','QUIC']
# Actually,in 'TLS' parts, there are two relevant protocols: TLSv1 and TLSv1.2,
# Because the number of TLSv1 is really small (<10), ignore them.
for i in range(len(protocols)):
    getTotal_StandAlone(protocols[i])
    getTotal_1v1(protocols[i])
    getTotal_5v5(protocols[i])



resultbook.save(result_file)
resultbook.close()
book.close()
book_1v1.close()
book_5v5.close()