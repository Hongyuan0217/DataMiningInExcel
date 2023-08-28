import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl import Workbook


result_file="/Users/guqingxian/Desktop/学习/Data Mining in Excel/Results_Exercises/results.xlsx"
book = load_workbook(result_file);

sheet=book.active;
sheet.title='Overall to Total'
book.create_sheet('Total 30 Traces')
book.create_sheet('Total 30 Traces Visualised')
book.create_sheet('Total 30 Traces Sorted')
book.create_sheet('Total HTTP')
book.create_sheet('HTTP Comparison')
book.create_sheet('Total DNS')
book.create_sheet('DNS Comparison')
book.create_sheet('Total MDNS')
book.create_sheet('MDNS Comparison')
book.create_sheet('Total LLMNR')
book.create_sheet('LLMNR Comparison')
book.create_sheet('Total NBNS')
book.create_sheet('NBNS Comparison')
book.create_sheet('Total SSDP')
book.create_sheet('SSDP Comparison')
book.create_sheet('Total NTP')
book.create_sheet('NTP Comparison')
book.create_sheet('Total SSL')
book.create_sheet('SSL Comparison')
book.create_sheet('Total TLS')
book.create_sheet('TLS Comparison')
book.create_sheet('Total TCP')
book.create_sheet('TCP Comparison')
book.create_sheet('Total UDP')
book.create_sheet('UDP Comparison')
book.create_sheet('Total QUIC')
book.create_sheet('QUIC Comparison')


book.save(result_file)
book.close()