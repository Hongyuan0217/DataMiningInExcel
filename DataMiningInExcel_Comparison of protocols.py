import pandas as pd
from openpyxl import load_workbook
import time


def findNo(column_index, target_string, trace, sheet):
    # trace:第n个出现的位置，例如 1 表示第一个出现的位置
    # column index: 列的数字索引，例如 3 表示第三列
    # 遍历指定列，找到第n个出现的目标字符串
    result = None
    count = 0
    # global startrow
    global Row

    # print('startrow=', startrow)
    for row in sheet.iter_rows(min_row=Row + 1, min_col=column_index, max_col=column_index, values_only=True):
        cell_value = row[0]
        Row += 1
        if cell_value == target_string:
            count += 1

            if count == trace:
                result = Row
                # print('Row=', Row)
                # startrow = Row+1
                # print('startrow==',startrow)
                break
    if result is not None:
        Row=0
        return result
    else:
        Row=0
        return sheet.max_row - 1

def getComparison(protocol) :
    resultsheet_name='{} Comparison'.format(protocol)
    resultsheet=resultbook[resultsheet_name]
    start_time = time.time()
    global total_length_SA, total_length_1v1, total_length_5v5
    cnt = 0
    length=0

    for row in datasheet.iter_rows(min_row=No26_StandAlone_startrow, max_row=No26_StandAlone_endrow-1, min_col=2,
                                   max_col=7, values_only=True):

        total_length_SA += int(row[4])
        if row[3] == '{}'.format(protocol):
            length += int(row[4])
            resultsheet.cell(row=cnt+5, column=2, value=row[4])
            resultsheet.cell(row=cnt+5, column=7, value=row[0])
            # print('filling the No.', cnt - 4, ' cell of {} Stand Alone'.format(protocol))
            cnt += 1

    endTime_No26 = time.time()
    print('The time of StandAlone {} of No.26 is'.format(protocol) , endTime_No26-start_time+LoadingTime_StandAlone
          ,' and the counts of data is ',cnt, ' and the bytes of data is',length)


    cnt = 0
    length=0

    for row in datasheet_1v1.iter_rows(min_row=No2_1v1_startrow, max_row=No2_1v1_endrow, min_col=5, max_col=10,
                                       values_only=True):

        total_length_1v1 += int(row[4])
        if row[3] == '{}'.format(protocol):
            length += int(row[4])
            resultsheet.cell(row=cnt+5, column=3, value=row[4])
            resultsheet.cell(row=cnt+5, column=8, value=row[0])
            # print('filling the No.', cnt - 4, ' cell of 1v1')
            cnt += 1

    endTime_No2 = time.time()
    print('The time of 1v1 {} of No.2 is'.format(protocol) , endTime_No2 - endTime_No26 + LoadingTime_1v1
          , ' and the counts of data is ', cnt, ' and the bytes of data is', length)


    cnt = 0
    length=0
    for row in datasheet_5v5.iter_rows(min_row=No17_5v5_startrow, max_row=No17_5v5_endrow, min_col=5, max_col=10,
                                       values_only=True):
        total_length_5v5 += int(row[4])
        if row[3] == '{}'.format(protocol):
            length += int(row[4])
            resultsheet.cell(row=cnt+5, column=4, value=row[4])
            resultsheet.cell(row=cnt+5, column=9, value=row[0])
            # print('filling the No.', cnt - 4, ' cell of 5v5')
            cnt += 1

    endTime_No17 = time.time()
    print('The time of 5v5 {} of No.17 is'.format(protocol) , endTime_No17 - endTime_No2 + LoadingTime_5v5
          ,' and the counts of data is ',cnt, ' and the bytes of data is',length)


result_file="/Users/guqingxian/Desktop/学习/Data Mining in Excel/Results_Exercises/results.xlsx"

StartTime_StandAlone=time.time()
book=load_workbook('/Users/guqingxian/Desktop/学习/Data Mining in Excel/Data Set/Stand-Alone.xlsx')
datasheet=book['单机 all30']
EndTime_StandAlone=time.time()
LoadingTime_StandAlone=EndTime_StandAlone-StartTime_StandAlone

StartTime_1v1=time.time()
book_1v1=load_workbook('/Users/guqingxian/Desktop/学习/Data Mining in Excel/Data Set/1v1.xlsx')
datasheet_1v1=book_1v1['backup']
EndTime_1v1=time.time()
LoadingTime_1v1=EndTime_1v1-StartTime_1v1


StartTime_5v5=time.time()
book_5v5=load_workbook('/Users/guqingxian/Desktop/学习/Data Mining in Excel/Data Set/5v5.xlsx')
datasheet_5v5=book_5v5['5v5']
EndTime_5v5=time.time()
LoadingTime_5v5=EndTime_5v5-StartTime_5v5

resultbook=load_workbook(result_file)

Row=0
No26_StandAlone_startrow=findNo(1,1,26,datasheet)
No26_StandAlone_endrow=findNo(1,1,27,datasheet)-1
No2_1v1_startrow=findNo(2,1,2,datasheet_1v1)
No2_1v1_endrow=findNo(2,1,3,datasheet_1v1)-1
No17_5v5_startrow=findNo(2,1,17,datasheet_5v5)
No17_5v5_endrow=findNo(2,1,18,datasheet_5v5)-1

total_length_SA=0
total_length_1v1=0
total_length_5v5=0

protocols=['HTTP','DNS','MDNS','LLMNR','NBNS','SSDP','NTP','SSL','TLSv1.2','TCP','UDP','QUIC']
# Actually,in 'TLS' parts, there are two relevant protocols: TLSv1 and TLSv1.2,
# Because the number of TLSv1 is really small (<10), ignore them.
for i in range(len(protocols)):
    getComparison(protocols[i])

print('the total length of StandAlone No.26 is ', total_length_SA)
print('the total length of 1v1 No.2 is ', total_length_1v1)
print('the total length of 5v5 No.17 is ', total_length_5v5)

resultbook.save(result_file)
resultbook.close()
book.close()
book_1v1.close()
book_5v5.close()