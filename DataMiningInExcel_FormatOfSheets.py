import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl import Workbook
from openpyxl.styles import Font

def TotalFormat(totalsheet,protocol):
    totalsheet.merge_cells('A1:D1')
    totalsheet['A1'] = '{} Data Length'.format(protocol)
    totalsheet['A1'].font = Font(bold=True)
    totalsheet['A1'].alignment = Alignment(horizontal='center', vertical='center')

    totalsheet.merge_cells('F1:I1')
    totalsheet['F1'] = '{} Delay'.format(protocol)
    totalsheet['F1'].font = Font(bold=True)
    totalsheet['F1'].alignment = Alignment(horizontal='center', vertical='center')

    totalsheet['A2'] = totalsheet['F2'] = 'Trace No.'
    totalsheet['B2'] = totalsheet['G2'] = 'Stand Alone'
    totalsheet['C2'] = totalsheet['H2'] = '1v1'
    totalsheet['D2'] = totalsheet['I2'] = '5v5'
    totalsheet.column_dimensions['B'].width = 18
    totalsheet.column_dimensions['G'].width = 18

    for i in range(30):
        totalsheet['A{}'.format(i + 3)] = i + 1
        totalsheet['F{}'.format(i + 3)] = i + 1

    totalsheet['A33'] = totalsheet['F33'] = 'Median'
    totalsheet['A34'] = totalsheet['F34'] = 'Average'

def ComparisonFormat(comparsionsheet,protocol):
    comparsionsheet.merge_cells('B1:I1')
    comparsionsheet['B1'] = 'Using Median and Average as the Representative Trace'
    comparsionsheet.merge_cells('B2:D2')
    comparsionsheet['B2'] = '{} Data Length'.format(protocol)
    comparsionsheet.merge_cells('G2:I2')
    comparsionsheet['G2'] = '{} Delay'.format(protocol)
    comparsionsheet['B3'] = comparsionsheet['G3'] = 'Stand Alone'
    comparsionsheet['C3'] = comparsionsheet['H3'] = '1v1'
    comparsionsheet['D3'] = comparsionsheet['I3'] = '5v5'
    comparsionsheet['A4'] = comparsionsheet['F4'] = 'Trace'
    comparsionsheet['B4'] = comparsionsheet['G4'] = 'No.26'
    comparsionsheet['C4'] = comparsionsheet['H4'] = 'No.2'
    comparsionsheet['D4'] = comparsionsheet['I4'] = 'No.17'



result_file="/Users/guqingxian/Desktop/学习/Data Mining in Excel/Results_Exercises/results.xlsx"
df=pd.read_excel(result_file)

# df.to_excel(result_file);
http_StandAlone=pd.read_excel("/Users/guqingxian/Desktop/学习/Data Mining in Excel/Data Set/Stand-Alone - HTTP.xlsx")
http_1v1 = pd.read_excel("/Users/guqingxian/Desktop/学习/Data Mining in Excel/Data Set/1v1 - HTTP.xlsx",usecols="F")
http_5v5 = pd.read_excel("/Users/guqingxian/Desktop/学习/Data Mining in Excel/Data Set/5v5 - HTTP.xlsx",usecols="F")



# s_HTTP_SA=pd.Series([http_StandAlone.max()[0],http_StandAlone.min()[0],(int)(http_StandAlone.mean()[0]),
#                      http_StandAlone.sum()[0],http_StandAlone.shape[0]],
#                     index=['max','min','avg','sum','package count'],name="HTTP-Stand Alone")
s_HTTP_1v1=pd.Series([http_1v1.max()[0],http_1v1.min()[0],http_1v1.mean()[0],http_1v1.sum()[0],http_1v1.shape[0]],
                    index=['max','min','avg','sum','package count'],name="HTTP-1v1")
s_HTTP_5v5=pd.Series([http_5v5.max()[0],http_5v5.min()[0],(int)(http_5v5.mean()[0]),http_5v5.sum()[0],http_5v5.shape[0]],
                    index=['max','min','avg','sum','package count'],name="HTTP-5v5")
# df=pd.DataFrame([s_HTTP_SA,s_HTTP_1v1,s_HTTP_5v5])
# print(df)

book = load_workbook(result_file);
sheet=book['Overall to Total']
Total30TracesSheet=book['Total 30 Traces']
TotalHTTP=book['Total HTTP']
TotalDNS=book['Total DNS']
TotalMDNS=book['Total MDNS']
TotalLLMNR=book['Total LLMNR']
TotalNBNS=book['Total NBNS']
TotalSSDP=book['Total SSDP']
TotalNTP=book['Total NTP']
TotalSSL=book['Total SSL']
TotalTLS=book['Total TLS']
TotalTCP=book['Total TCP']
TotalUDP=book['Total UDP']
TotalQUIC=book['Total QUIC']

HTTPComparison=book['HTTP Comparison']
DNSComparison=book['DNS Comparison']
MDNSComparison=book['MDNS Comparison']
LLMNRComparison=book['LLMNR Comparison']
NBNSComparison=book['NBNS Comparison']
SSDPComparison=book['SSDP Comparison']
NTPComparison=book['NTP Comparison']
SSLComparison=book['SSL Comparison']
TLSComparison=book['TLS Comparison']
TCPComparison=book['TCP Comparison']
UDPComparison=book['UDP Comparison']
QUICComparison=book['QUIC Comparison']



#design excel format - Overall to Total
ways=['HTTP','DNS','MDNS','LLMNR','NBNS','SSDP','NTP','SSL','TLS','TCP','UDP','QUIC']
types=['Stand Alone','1v1','5v5']
sheet.merge_cells('A1:A2')
sheet.merge_cells('B1:B2')
sheet.merge_cells('C1:F1')
sheet['C1']="Packet Size (Bytes)"
sheet['C1'].alignment = Alignment(horizontal='center', vertical='center')
sheet['C2']='Max'
sheet['D2']='Min'
sheet['E2']='Avg'
sheet['F2']='Sum'
sheet.merge_cells('G1:G2')
sheet['G1']='Package ' \
            'Count'
sheet.column_dimensions['B'].width = 18
sheet.column_dimensions['G'].width = 18
sheet['G1'].alignment = Alignment(horizontal='center', vertical='center')
for i in range(len(ways)):
    start_row = 3 * (i + 1)
    end_row = start_row + 2
    merge_range = 'A{}:A{}'.format(start_row, end_row)

    top_left_cell = 'A{}'.format(start_row)
    sheet[top_left_cell] = ways[i]

    sheet.merge_cells(merge_range)
    sheet['A{}'.format(start_row)].alignment = Alignment(horizontal='center', vertical='center')

    sheet['B{}'.format(3*(i+1))]=types[0];
    sheet['B{}'.format(3*(i+1)+1)]=types[1];
    sheet['B{}'.format(3*(i+1)+2)]=types[2];

#design excel format - Total 30 Traces Sheet
Total30TracesSheet['A1']='Trace'
Total30TracesSheet['B1']='Stand Alone'
Total30TracesSheet['C1']='Stand Alone'
Total30TracesSheet['D1']='1v1'
Total30TracesSheet['E1']='1v1'
Total30TracesSheet['F1']='5v5'
Total30TracesSheet['G1']='5v5'

Total30TracesSheet['A2']='No.'
Total30TracesSheet['B2']='Duration'
Total30TracesSheet['C2']='Packets'
Total30TracesSheet['D2']='Duration'
Total30TracesSheet['E2']='Packets'
Total30TracesSheet['F2']='Duration'
Total30TracesSheet['G2']='Packets'

Total30TracesSheet['A1'].font=Font(bold=True)
Total30TracesSheet['B1'].font=Font(bold=True)
Total30TracesSheet['C1'].font=Font(bold=True)
Total30TracesSheet['D1'].font=Font(bold=True)
Total30TracesSheet['E1'].font=Font(bold=True)
Total30TracesSheet['F1'].font=Font(bold=True)


TotalFormat(TotalHTTP,'HTTP')
TotalFormat(TotalDNS,'DNS')
TotalFormat(TotalMDNS,'MDNS')
TotalFormat(TotalLLMNR,'LLMNR')
TotalFormat(TotalNBNS,'NBNS')
TotalFormat(TotalSSDP,'SSDP')
TotalFormat(TotalNTP,'NTP')
TotalFormat(TotalSSL,'SSL')
TotalFormat(TotalTLS,'TLS')
TotalFormat(TotalTCP,'TCP')
TotalFormat(TotalUDP,'UDP')
TotalFormat(TotalQUIC,'QUIC')

ComparisonFormat(HTTPComparison,'HTTP')
ComparisonFormat(DNSComparison,'DNS')
ComparisonFormat(MDNSComparison,'MDNS')
ComparisonFormat(LLMNRComparison,'LLMNR')
ComparisonFormat(NBNSComparison,'NBNS')
ComparisonFormat(SSDPComparison,'SSDP')
ComparisonFormat(NTPComparison,'NTP')
ComparisonFormat(SSLComparison,'SSL')
ComparisonFormat(TLSComparison,'TLS')
ComparisonFormat(TCPComparison,'TCP')
ComparisonFormat(UDPComparison,'UDP')
ComparisonFormat(QUICComparison,'QUIC')





# #design excel format - Total HTTP
# TotalHTTP.merge_cells('A1:D1')
# TotalHTTP['A1']='HTTP Data Length'
# TotalHTTP['A1'].font=Font(bold=True)
# TotalHTTP['A1'].alignment = Alignment(horizontal='center', vertical='center')
#
# TotalHTTP.merge_cells('F1:I1')
# TotalHTTP['F1']='HTTP Delay'
# TotalHTTP['F1'].font=Font(bold=True)
# TotalHTTP['F1'].alignment = Alignment(horizontal='center', vertical='center')
#
# TotalHTTP['A2']=TotalHTTP['F2']='Trace No.'
# TotalHTTP['B2']=TotalHTTP['G2']='Stand Alone'
# TotalHTTP['C2']=TotalHTTP['H2']='1v1'
# TotalHTTP['D2']=TotalHTTP['I2']='5v5'
# TotalHTTP.column_dimensions['B'].width = 18
# TotalHTTP.column_dimensions['G'].width = 18
#
# for i in range(30):
#     TotalHTTP['A{}'.format(i+3)]=i+1
#     TotalHTTP['F{}'.format(i + 3)] = i + 1
#
# TotalHTTP['A33']=TotalHTTP['F33']='Median'
# TotalHTTP['A34']=TotalHTTP['F34']='Average'
#
#
# #design excel format - Total TCP
# TotalTCP.merge_cells('A1:D1')
# TotalTCP['A1']='TCP Data Length'
# TotalTCP['A1'].font=Font(bold=True)
# TotalTCP['A1'].alignment = Alignment(horizontal='center', vertical='center')
#
# TotalTCP.merge_cells('F1:I1')
# TotalTCP['F1']='TCP Delay'
# TotalTCP['F1'].font=Font(bold=True)
# TotalTCP['F1'].alignment = Alignment(horizontal='center', vertical='center')
#
# TotalTCP['A2']=TotalTCP['F2']='Trace No.'
# TotalTCP['B2']=TotalTCP['G2']='Stand Alone'
# TotalTCP['C2']=TotalTCP['H2']='1v1'
# TotalTCP['D2']=TotalTCP['I2']='5v5'
# TotalTCP.column_dimensions['B'].width = 18
# TotalTCP.column_dimensions['G'].width = 18
#
# for i in range(30):
#     TotalTCP['A{}'.format(i+3)]=i+1
#     TotalTCP['F{}'.format(i + 3)] = i + 1
#
# TotalTCP['A33']=TotalTCP['F33']='Median'
# TotalTCP['A34']=TotalTCP['F34']='Average'
#
# #design excel format - Total UDP
# TotalUDP.merge_cells('A1:D1')
# TotalUDP['A1']='UDP Data Length'
# TotalUDP['A1'].font=Font(bold=True)
# TotalUDP['A1'].alignment = Alignment(horizontal='center', vertical='center')
#
# TotalUDP.merge_cells('F1:I1')
# TotalUDP['F1']='UDP Delay'
# TotalUDP['F1'].font=Font(bold=True)
# TotalUDP['F1'].alignment = Alignment(horizontal='center', vertical='center')
#
# TotalUDP['A2']=TotalTCP['F2']='Trace No.'
# TotalUDP['B2']=TotalTCP['G2']='Stand Alone'
# TotalUDP['C2']=TotalTCP['H2']='1v1'
# TotalUDP['D2']=TotalTCP['I2']='5v5'
# TotalUDP.column_dimensions['B'].width = 18
# TotalUDP.column_dimensions['G'].width = 18
#
# for i in range(30):
#     TotalUDP['A{}'.format(i+3)]=i+1
#     TotalUDP['F{}'.format(i + 3)] = i + 1
#
# TotalUDP['A33']=TotalUDP['F33']='Median'
# TotalUDP['A34']=TotalUDP['F34']='Average'
#
# #design excel format - Total QUIC
# TotalQUIC.merge_cells('A1:D1')
# TotalQUIC['A1']='QUIC Data Length'
# TotalQUIC['A1'].font=Font(bold=True)
# TotalQUIC['A1'].alignment = Alignment(horizontal='center', vertical='center')
#
# TotalQUIC.merge_cells('F1:I1')
# TotalQUIC['F1']='QUIC Delay'
# TotalQUIC['F1'].font=Font(bold=True)
# TotalQUIC['F1'].alignment = Alignment(horizontal='center', vertical='center')
#
# TotalQUIC['A2']=TotalQUIC['F2']='Trace No.'
# TotalQUIC['B2']=TotalQUIC['G2']='Stand Alone'
# TotalQUIC['C2']=TotalQUIC['H2']='1v1'
# TotalQUIC['D2']=TotalQUIC['I2']='5v5'
# TotalQUIC.column_dimensions['B'].width = 18
# TotalQUIC.column_dimensions['G'].width = 18
#
# for i in range(30):
#     TotalQUIC['A{}'.format(i+3)]=i+1
#     TotalQUIC['F{}'.format(i + 3)] = i + 1
#
# TotalQUIC['A33']=TotalQUIC['F33']='Median'
# TotalQUIC['A34']=TotalQUIC['F34']='Average'
#
# #design excel format - HTTP Comparision
# HTTPComparision.merge_cells('B1:I1')
# HTTPComparision['B1']='Using Median and Average as the Representative Trace'
# HTTPComparision.merge_cells('B2:D2')
# HTTPComparision['B2']='HTTP Data Length'
# HTTPComparision.merge_cells('G2:I2')
# HTTPComparision['G2']='HTTP Delay'
# HTTPComparision['B3']=HTTPComparision['G3']='Stand Alone'
# HTTPComparision['C3']=HTTPComparision['H3']='1v1'
# HTTPComparision['D3']=HTTPComparision['I3']='5v5'
# HTTPComparision['A4']=HTTPComparision['F4']='Trace'
# HTTPComparision['B4']=HTTPComparision['G4']='No.26'
# HTTPComparision['C4']=HTTPComparision['H4']='No.2'
# HTTPComparision['D4']=HTTPComparision['I4']='No.17'
#
# #design excel format - TCP Comparision
# TCPComparision.merge_cells('B1:I1')
# TCPComparision['B1']='Using Median and Average as the Representative Trace'
# TCPComparision.merge_cells('B2:D2')
# TCPComparision['B2']='TCP Data Length'
# TCPComparision.merge_cells('G2:I2')
# TCPComparision['G2']='TCP Delay'
# TCPComparision['B3']=TCPComparision['G3']='Stand Alone'
# TCPComparision['C3']=TCPComparision['H3']='1v1'
# TCPComparision['D3']=TCPComparision['I3']='5v5'
# TCPComparision['A4']=TCPComparision['F4']='Trace'
# TCPComparision['B4']=TCPComparision['G4']='No.26'
# TCPComparision['C4']=TCPComparision['H4']='No.2'
# TCPComparision['D4']=TCPComparision['I4']='No.17'
#
# #design excel format - UDP Comparision
# UDPComparision.merge_cells('B1:I1')
# UDPComparision['B1']='Using Median and Average as the Representative Trace'
# UDPComparision.merge_cells('B2:D2')
# UDPComparision['B2']='UDP Data Length'
# UDPComparision.merge_cells('G2:I2')
# UDPComparision['G2']='UDP Delay'
# UDPComparision['B3']=UDPComparision['G3']='Stand Alone'
# UDPComparision['C3']=UDPComparision['H3']='1v1'
# UDPComparision['D3']=UDPComparision['I3']='5v5'
# UDPComparision['A4']=UDPComparision['F4']='Trace'
# UDPComparision['B4']=UDPComparision['G4']='No.26'
# UDPComparision['C4']=UDPComparision['H4']='No.2'
# UDPComparision['D4']=UDPComparision['I4']='No.17'
#
# #design excel format - QUIC Comparision
# QUICComparision.merge_cells('B1:I1')
# QUICComparision['B1']='Using Median and Average as the Representative Trace'
# QUICComparision.merge_cells('B2:D2')
# QUICComparision['B2']='QUIC Data Length'
# QUICComparision.merge_cells('G2:I2')
# QUICComparision['G2']='QUIC Delay'
# QUICComparision['B3']=QUICComparision['G3']='Stand Alone'
# QUICComparision['C3']=QUICComparision['H3']='1v1'
# QUICComparision['D3']=QUICComparision['I3']='5v5'
# QUICComparision['A4']=QUICComparision['F4']='Trace'
# QUICComparision['B4']=QUICComparision['G4']='No.26'
# QUICComparision['C4']=QUICComparision['H4']='No.2'
# QUICComparision['D4']=QUICComparision['I4']='No.17'





book.save(result_file);
book.close()