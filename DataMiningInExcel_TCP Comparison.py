import pandas as pd
from openpyxl import load_workbook
import time

start_time=time.time()
Row=0
def findNo(column_index, target_string, trace, sheet):
    # trace:第n个出现的位置，例如 1 表示第一个出现的位置
    # column index: 列的数字索引，例如 3 表示第三列
    # 遍历指定列，找到第n个出现的目标字符串
    result = None
    count = 0
    # global startrow
    global Row

    # print('startrow=', startrow)
    for row in sheet.iter_rows(min_row=Row + 1, min_col=column_index, max_col=column_index, values_only=True):
        cell_value = row[0]
        Row += 1
        if cell_value == target_string:
            count += 1

            if count == trace:
                result = Row
                # print('Row=', Row)
                # startrow = Row+1
                # print('startrow==',startrow)
                break
    if result is not None:
        Row=0
        return result
    else:
        Row=0
        return sheet.max_row - 1


result_file="/Users/guqingxian/Desktop/学习/Data Mining in Excel/Results_Exercises/results.xlsx"
StandAlone=pd.read_excel("/Users/guqingxian/Desktop/学习/Data Mining in Excel/Data Set/Stand-Alone.xlsx")
book=load_workbook('/Users/guqingxian/Desktop/学习/Data Mining in Excel/Data Set/Stand-Alone.xlsx')
datasheet=book['单机 all30']
book_1v1=load_workbook('/Users/guqingxian/Desktop/学习/Data Mining in Excel/Data Set/1v1.xlsx')
datasheet_1v1=book_1v1['backup']
book_5v5=load_workbook('/Users/guqingxian/Desktop/学习/Data Mining in Excel/Data Set/5v5.xlsx')
datasheet_5v5=book_5v5['5v5']

resultbook=load_workbook(result_file)
resultsheet=resultbook['TCP Comparison']

No26_StandAlone_startrow=findNo(1,1,26,datasheet)
No26_StandAlone_endrow=findNo(1,1,27,datasheet)-1
No2_1v1_startrow=findNo(2,1,2,datasheet_1v1)
No2_1v1_endrow=findNo(2,1,3,datasheet_1v1)-1
No17_5v5_startrow=findNo(2,1,17,datasheet_5v5)
No17_5v5_endrow=findNo(2,1,18,datasheet_5v5)-1

cnt=5
for row in datasheet.iter_rows(min_row=No26_StandAlone_startrow,max_row=No26_StandAlone_endrow, min_col=2, max_col=6, values_only=True):

    if row[3] == 'TCP':
        resultsheet.cell(row=cnt,column=2,value=row[4])
        resultsheet.cell(row=cnt,column=7,value=row[0])
        print('filling the No.',cnt-4,' cell of Stand Alone')
        cnt += 1



cnt = 5
for row in datasheet_1v1.iter_rows(min_row=No2_1v1_startrow,max_row=No2_1v1_endrow, min_col=5, max_col=9, values_only=True):

    if row[3] == 'TCP':
        resultsheet.cell(row=cnt,column=3,value=row[4])
        resultsheet.cell(row=cnt,column=8,value=row[0])
        print('filling the No.', cnt - 4, ' cell of 1v1')
        cnt += 1

cnt=5
for row in datasheet_5v5.iter_rows(min_row=No17_5v5_startrow,max_row=No17_5v5_endrow, min_col=5, max_col=9, values_only=True):
    if row[3] == 'TCP':
        resultsheet.cell(row=cnt,column=4,value=row[4])
        resultsheet.cell(row=cnt,column=9,value=row[0])
        print('filling the No.', cnt - 4, ' cell of 5v5')
        cnt += 1

print('before 1v1 load_workbook')
book_1v1=load_workbook('/Users/guqingxian/Desktop/学习/Data Mining in Excel/Data Set/1v1.xlsx')
datasheet_1v1=book_1v1['backup']
print('after 1v1 load_wordbook')
traces_1v1=[]
for i in range(30):
    traces_1v1.append(findNo(2,1,1,datasheet_1v1))
Row=0


resultbook.save(result_file)
resultbook.close()
book.close()
book_1v1.close()
book_5v5.close()

end_time=time.time()
print('total time is ',end_time-start_time)
